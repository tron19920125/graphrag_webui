import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO


def render_excel_file(uploaded_file): 
    wb = load_workbook(uploaded_file)

    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    even_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    odd_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    alignment = Alignment(wrap_text=True)
    header_alignment = Alignment(wrap_text=False)
    thin_side = Side(border_style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    dv = DataValidation(type="list", formula1='"是,否,勉强通过"', allow_blank=True)

    for sheet in wb.worksheets:
        max_row = sheet.max_row
        max_col = sheet.max_column

        headers = ["是否通过", "不通过原因", "通过但仍需改进"]
        
        for i, header_text in enumerate(headers, start=max_col + 1):
            sheet.cell(row=1, column=i, value=header_text)
        
        max_col = sheet.max_column

        for col in range(1, max_col + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 25

        pass_col_letter = get_column_letter(max_col - 2)

        dv.add(f"{pass_col_letter}2:{pass_col_letter}{max_row}")
        sheet.add_data_validation(dv)

        for r, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
            for cell in row:
                cell.border = border
                if r == 1:
                    cell.alignment = header_alignment
                    cell.fill = header_fill
                else:
                    cell.alignment = alignment
                    if r % 2 == 0:
                        cell.fill = even_fill
                    else:
                        cell.fill = odd_fill

        for r in range(2, max_row + 1):
            sheet.cell(row=r, column=max_col - 2, value="-")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
